import argparse

import openpyxl
from datetime import datetime
from dataclasses import dataclass

SALES_PER_PAGE = 7
EXCLUDE_ORDER_TYPES = ["RS STC"]

# Export as yearly/daily USD/SEK as xlsx from
# https://www.riksbank.se/sv/statistik/rantor-och-valutakurser/sok-rantor-och-valutakurser

# It is allowed to use either the daily exchange rate or the yearly exchange rate,
# as long as it is consistent for all trades.
YEARLY_EXCHANGE_RATE_FILE = "usd_sek_yearly.xlsx"
DAILY_EXCHANGE_RATE_FILE = "usd_sek_daily.xlsx"

YEARLY_EXCHANGE_RATE_PARAM = "year"
DAILY_EXCHANGE_RATE_PARAM = "day"

parser = argparse.ArgumentParser(prog="k4", description='Generate K4 SRU files.')
parser.add_argument('--org-nummer', required=True,
                    help='org eller personnummer')
parser.add_argument('--fullt-namn', required=True,
                    help='för och efternamn')
parser.add_argument('--adress', required=True,
                    help='adress')
parser.add_argument('--postnummer', required=True,
                    help='postnummer')
parser.add_argument('--postort', required=True,
                    help='postort')
parser.add_argument('--epost', required=True,
                    help='epost')
parser.add_argument('--year', required=True,
                    help='vilket räkenskapsår')
parser.add_argument('--rate', required=True,
                    help=f'{YEARLY_EXCHANGE_RATE_PARAM} or {DAILY_EXCHANGE_RATE_PARAM}')
parser.add_argument('--trades', required=True,
                    help='affärer i xlsx')


@dataclass
class Trade:
  antal: int
  beteckning: str
  forsaljsningspris: float
  omkostnadsbelopp: float
  gainloss: float
  ordertype: str
  date: datetime

  def vinst(self):
    return self.gainloss if self.gainloss > 0 else 0

  def forlust(self):
    return -self.gainloss if self.gainloss < 0 else 0


def parse_trades(file):
  trades = openpyxl.load_workbook(file)
  for row in trades.active.rows:
    if row[0].value == 'Sell':
      yield Trade(antal=row[3].value,
                  beteckning=row[1].value,
                  omkostnadsbelopp=row[10].value,
                  forsaljsningspris=row[13].value,
                  gainloss=row[18].value,
                  ordertype=row[28].value if len(row) > 28 else None,
                  date=datetime.strptime(row[12].value, "%m/%d/%Y"))


def read_yearly_exchange_rate(year):
  workbook = openpyxl.load_workbook(YEARLY_EXCHANGE_RATE_FILE)
  sheet = workbook.active
  for row in sheet.rows:
    if str(row[0].value) == str(year):
      return float(row[1].value)

  raise ValueError(f"No exchange rate found for year {year}")

def read_daily_exchange_rate(date):
  workbook = openpyxl.load_workbook(DAILY_EXCHANGE_RATE_FILE)
  sheet = workbook.active
  exchange_rates = {}
  for row in sheet.rows:
    try:
      exchange_date = datetime.strptime(str(row[0].value), "%Y-%m-%d")
    except ValueError:
      continue
    exchange_rate = float(row[1].value)
    exchange_rates[exchange_date] = exchange_rate

  if date in exchange_rates:
    return exchange_rates[date]

  closest_date = min(exchange_rates.keys(), key=lambda d: abs(d - date))
  if abs((closest_date - date).days) > 7:
    raise ValueError(f"No exchange rate found within seven days of {date}")
  return exchange_rates[closest_date]

def main():
  # https://www.skatteverket.se/foretag/etjansterochblanketter/blanketterbroschyrer/broschyrer/info/269.4.39f16f103821c58f680007305.html
  args = parser.parse_args()
  with open("INFO.SRU", mode="w") as outfile:
    outfile.write("#DATABESKRIVNING_START\n")
    outfile.write("#PRODUKT SRU\n")
    outfile.write("#FILNAMN BLANKETTER.SRU\n")
    outfile.write("#DATABESKRIVNING_SLUT\n")
    outfile.write("#MEDIELEV_START\n")
    outfile.write("#ORGNR " + args.org_nummer + "\n")
    outfile.write("#NAMN " + args.fullt_namn + "\n")
    outfile.write("#ADRESS " + args.adress + "\n")
    outfile.write("#POSTNR " + args.postnummer + "\n")
    outfile.write("#POSTORT " + args.postort + "\n")
    outfile.write("#EMAIL " + args.epost + "\n")
    outfile.write("#MEDIELEV_SLUT\n")

  trades = [trade for trade in list(parse_trades(args.trades)) if trade.ordertype not in EXCLUDE_ORDER_TYPES]
  now = datetime.now()

  total_vinst_sek = 0;
  total_forlust_sek = 0;

  with open("BLANKETTER.SRU", mode="w") as outfile:
    chunk_counter = 1
    for chunk in chunk_list(trades, SALES_PER_PAGE):
      counter = 10
      outfile.write("#BLANKETT K4-" + args.year + "P4\n")
      outfile.write("#IDENTITET " + args.org_nummer + " " + now.strftime('%Y%m%d %H%M%S') + "\n")
      outfile.write("#NAMN " + args.fullt_namn + "\n")
      for trade in chunk:
        if counter - 31 > SALES_PER_PAGE:
          raise ValueError("Can only have %d sales per page" % SALES_PER_PAGE)

        if args.rate == YEARLY_EXCHANGE_RATE_PARAM:
          exchange_rate = read_yearly_exchange_rate(args.year)
        elif args.rate == DAILY_EXCHANGE_RATE_PARAM:
          exchange_rate = read_daily_exchange_rate(trade.date)
        else:
          raise ValueError("Correct rate must be supplied")

        vinst_sek = trade.vinst() * exchange_rate
        forlust_sek = trade.forlust() * exchange_rate
        total_vinst_sek += vinst_sek
        total_forlust_sek += forlust_sek

        outfile.write("#UPPGIFT 3" + str(counter) + "0 " + str(trade.antal) + "\n")
        outfile.write("#UPPGIFT 3" + str(counter) + "1 " + trade.beteckning + "\n")
        outfile.write("#UPPGIFT 3" + str(counter) + "2 " + str(round(trade.forsaljsningspris * exchange_rate)) + "\n")
        outfile.write("#UPPGIFT 3" + str(counter) + "3 " + str(round(trade.omkostnadsbelopp * exchange_rate)) + "\n")
        outfile.write("#UPPGIFT 3" + str(counter) + "4 " + str(round(vinst_sek)) + "\n")
        outfile.write("#UPPGIFT 3" + str(counter) + "5 " + str(round(forlust_sek)) + "\n")
        counter += 1
      outfile.write("#UPPGIFT 7014 " + str(chunk_counter) + "\n")
      outfile.write("#BLANKETTSLUT\n")
      chunk_counter += 1
    outfile.write("#FIL_SLUT\n")

  print("K4 SRU files generated. Year: {0}, Rate: {1}".format(args.year, args.rate))
  print("Summa vinst SEK: " + swedish_float(total_vinst_sek))
  print("Summa förlust SEK: " + swedish_float(total_forlust_sek))
  print("Nettovinst/Nettoförlust SEK: " + swedish_float(total_vinst_sek - total_forlust_sek))



def chunk_list(lst, size):
  return [lst[i:i + size] for i in range(0, len(lst), size)]


def swedish_float(f):
  return str(f).replace('.', ',')


if __name__ == "__main__":
  main()
